"""
Excel Exporter for Audit Reports

Note: Uses openpyxl if available, falls back to CSV with .xlsx extension
"""

from pathlib import Path
from typing import Dict, Any

from .base_exporter import BaseExporter, logger


class ExcelExporter(BaseExporter):
    """Export audit reports as Excel"""
    
    def __init__(self):
        self.openpyxl_available = False
        try:
            import openpyxl
            self.openpyxl_available = True
            self.openpyxl = openpyxl
        except ImportError:
            logger.warning("openpyxl not available, using fallback Excel generation")
    
    async def export(self, audit_result: Dict[str, Any], output_file: Path):
        """Export single audit result as Excel"""
        self._ensure_output_dir(output_file)
        
        if self.openpyxl_available:
            await self._export_openpyxl(audit_result, output_file)
        else:
            await self._export_fallback(audit_result, output_file)
        
        logger.debug(f"Exported Excel to {output_file}")
    
    async def export_batch_summary(self, batch_results: Dict[str, Any], output_file: Path):
        """Export batch summary as Excel"""
        self._ensure_output_dir(output_file)
        
        if self.openpyxl_available:
            await self._export_batch_openpyxl(batch_results, output_file)
        else:
            await self._export_batch_fallback(batch_results, output_file)
        
        logger.debug(f"Exported batch Excel to {output_file}")
    
    async def _export_openpyxl(self, audit_result: Dict[str, Any], output_file: Path):
        """Export using openpyxl library"""
        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Report"
        
        # Header
        contract_name = audit_result.get('contract_name', 'Unknown')
        severity = audit_result.get('severity', 'unknown')
        
        ws['A1'] = 'Contract Name'
        ws['B1'] = contract_name
        ws['A2'] = 'Severity'
        ws['B2'] = severity.upper()
        ws['A3'] = 'Total Findings'
        ws['B3'] = len(audit_result.get('findings', []))
        
        # Findings table
        ws['A5'] = '#'
        ws['B5'] = 'Title'
        ws['C5'] = 'Severity'
        ws['D5'] = 'Type'
        ws['E5'] = 'Description'
        ws['F5'] = 'Location'
        ws['G5'] = 'Recommendation'
        
        # Bold headers
        for cell in ws[5]:
            cell.font = self.openpyxl.styles.Font(bold=True)
        
        # Findings data
        findings = audit_result.get('findings', [])
        for idx, finding in enumerate(findings, 1):
            row = 5 + idx
            ws[f'A{row}'] = idx
            ws[f'B{row}'] = finding.get('title', '')
            ws[f'C{row}'] = finding.get('severity', '')
            ws[f'D{row}'] = finding.get('type', '')
            ws[f'E{row}'] = finding.get('description', '')
            ws[f'F{row}'] = finding.get('location', '')
            ws[f'G{row}'] = finding.get('recommendation', '')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 50
        
        wb.save(output_file)
    
    async def _export_batch_openpyxl(self, batch_results: Dict[str, Any], output_file: Path):
        """Export batch summary using openpyxl"""
        wb = self.openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary = batch_results.get('summary', {})
        
        ws_summary['A1'] = 'Metric'
        ws_summary['B1'] = 'Value'
        ws_summary['A1'].font = self.openpyxl.styles.Font(bold=True)
        ws_summary['B1'].font = self.openpyxl.styles.Font(bold=True)
        
        metrics = [
            ('Total Contracts', batch_results.get('total_contracts', 0)),
            ('Successful', batch_results.get('successful', 0)),
            ('Failed', batch_results.get('failed', 0)),
            ('Risk Score', f"{summary.get('risk_score', 0)}/100"),
            ('Total Findings', summary.get('total_findings', 0)),
            ('Avg Findings/Contract', summary.get('avg_findings_per_contract', 0))
        ]
        
        for idx, (metric, value) in enumerate(metrics, 2):
            ws_summary[f'A{idx}'] = metric
            ws_summary[f'B{idx}'] = value
        
        # Risk distribution
        risk_dist = summary.get('risk_distribution', {})
        row = len(metrics) + 3
        ws_summary[f'A{row}'] = 'Severity'
        ws_summary[f'B{row}'] = 'Count'
        ws_summary[f'A{row}'].font = self.openpyxl.styles.Font(bold=True)
        ws_summary[f'B{row}'].font = self.openpyxl.styles.Font(bold=True)
        
        for idx, (severity, count) in enumerate(risk_dist.items(), 1):
            ws_summary[f'A{row + idx}'] = severity.title()
            ws_summary[f'B{row + idx}'] = count
        
        # Contracts sheet
        ws_contracts = wb.create_sheet("Contracts")
        ws_contracts['A1'] = 'Contract Name'
        ws_contracts['B1'] = 'Severity'
        ws_contracts['C1'] = 'Findings'
        ws_contracts['D1'] = 'Status'
        
        for cell in ws_contracts[1]:
            cell.font = self.openpyxl.styles.Font(bold=True)
        
        contracts = batch_results.get('contracts', [])
        for idx, contract in enumerate(contracts, 2):
            name = contract.get('contract_name', 'Unknown')
            ws_contracts[f'A{idx}'] = name
            
            if contract.get('success'):
                severity = contract.get('severity', 'unknown')
                findings_count = len(contract.get('findings', []))
                ws_contracts[f'B{idx}'] = severity.upper()
                ws_contracts[f'C{idx}'] = findings_count
                ws_contracts[f'D{idx}'] = 'Success'
            else:
                ws_contracts[f'B{idx}'] = 'N/A'
                ws_contracts[f'C{idx}'] = 'N/A'
                ws_contracts[f'D{idx}'] = 'Failed'
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_contracts.column_dimensions['A'].width = 30
        ws_contracts.column_dimensions['B'].width = 15
        ws_contracts.column_dimensions['C'].width = 12
        ws_contracts.column_dimensions['D'].width = 12
        
        wb.save(output_file)
    
    async def _export_fallback(self, audit_result: Dict[str, Any], output_file: Path):
        """Fallback: Export as CSV with .xlsx extension"""
        import csv
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            contract_name = audit_result.get('contract_name', 'Unknown')
            severity = audit_result.get('severity', 'unknown')
            
            writer.writerow(['Contract Name', contract_name])
            writer.writerow(['Severity', severity.upper()])
            writer.writerow([])
            
            writer.writerow(['#', 'Title', 'Severity', 'Type', 'Description'])
            
            findings = audit_result.get('findings', [])
            for idx, finding in enumerate(findings, 1):
                writer.writerow([
                    idx,
                    finding.get('title', ''),
                    finding.get('severity', ''),
                    finding.get('type', ''),
                    finding.get('description', '')
                ])
            
            writer.writerow([])
            writer.writerow(['Note: Install openpyxl for full Excel support'])
    
    async def _export_batch_fallback(self, batch_results: Dict[str, Any], output_file: Path):
        """Fallback: Export batch as CSV"""
        import csv
        
        summary = batch_results.get('summary', {})
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            writer.writerow(['Batch Audit Summary'])
            writer.writerow([])
            writer.writerow(['Total Contracts', batch_results.get('total_contracts', 0)])
            writer.writerow(['Successful', batch_results.get('successful', 0)])
            writer.writerow(['Failed', batch_results.get('failed', 0)])
            writer.writerow(['Risk Score', f"{summary.get('risk_score', 0)}/100"])
            writer.writerow([])
            
            writer.writerow(['Contract Name', 'Severity', 'Findings', 'Status'])
            
            contracts = batch_results.get('contracts', [])
            for contract in contracts:
                name = contract.get('contract_name', 'Unknown')
                if contract.get('success'):
                    severity = contract.get('severity', 'unknown')
                    findings_count = len(contract.get('findings', []))
                    writer.writerow([name, severity.upper(), findings_count, 'Success'])
                else:
                    writer.writerow([name, 'N/A', 'N/A', 'Failed'])
            
            writer.writerow([])
            writer.writerow(['Note: Install openpyxl for full Excel support'])

